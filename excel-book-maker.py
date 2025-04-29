import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

# Sample data
data = [
    {"name": "Amesha", "type": "Archwing", "category": "Archwing", "obtained": False, "max_rank": False},
    {"name": "Ash", "type": "Warframe", "category": "Warframes", "obtained": True, "max_rank": True},
    {"name": "Ash Prime", "type": "Warframe", "category": "Warframes", "obtained": True, "max_rank": True},
    {"name": "Atlas", "type": "Warframe", "category": "Warframes", "obtained": False, "max_rank": False},
    {"name": "Atlas Prime", "type": "Warframe", "category": "Warframes", "obtained": False, "max_rank": False},
    {"name": "Banshee", "type": "Warframe", "category": "Warframes", "obtained": True, "max_rank": True},
    {"name": "Banshee Prime", "type": "Warframe", "category": "Warframes", "obtained": True, "max_rank": True},
    {"name": "Baruuk", "type": "Warframe", "category": "Warframes", "obtained": False, "max_rank": False},
    {"name": "Baruuk Prime", "type": "Warframe", "category": "Warframes", "obtained": False, "max_rank": False},
    {"name": "Bonewidow", "type": "Warframe", "category": "Warframes", "obtained": False, "max_rank": False}
]

df = pd.DataFrame(data)

# Add progress column
def get_progress_symbol(row):
    if not row["obtained"]:
        return "❌"
    elif row["obtained"] and not row["max_rank"]:
        return "🔵"
    elif row["obtained"] and row["max_rank"]:
        return "✅"

df["progress"] = df.apply(get_progress_symbol, axis=1)

# Create a workbook
wb = Workbook()
ws = wb.active
ws.title = "Progress Tracker"

# Add DataFrame to worksheet
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    ws.append(row)
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        # Header style
        if r_idx == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        # Center align progress column
        if df.columns[c_idx - 1] == "progress":
            cell.alignment = Alignment(horizontal='center')
            if value == "✅":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
            elif value == "🔵":
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
            elif value == "❌":
                cell.fill = PatternFill(start_color="E2E2E2", end_color="E2E2E2", fill_type="solid")  # light gray

# Create an Excel Table
table = Table(displayName="WarframeProgress", ref=f"A1:F{len(df)+1}")
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# Save the file
wb.save("warframe_beautiful_tracker.xlsx")
